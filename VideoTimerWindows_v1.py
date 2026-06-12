import sys
import os
import time
import cv2
import json
import csv
import numpy as np
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QPushButton, QVBoxLayout, QLabel,
    QFileDialog, QHBoxLayout, QMessageBox, QSlider, QComboBox, QDialog,
    QLineEdit, QFormLayout, QListWidget, QSizePolicy, QInputDialog, QAction,
    QCheckBox, QGroupBox, QColorDialog, QTableWidget, QTableWidgetItem,
    QHeaderView, QDialogButtonBox, QSpacerItem, QSizePolicy as QSzPolicy
)
from PyQt5.QtCore import (
    Qt, QTimer, QThread, pyqtSignal, QEvent, QRectF,
    QMutex, QMutexLocker, pyqtSlot, QSettings
)
from PyQt5.QtGui import QPalette, QColor, QPainter, QImage, QKeySequence, QPen

logging.basicConfig(
    filename='behaviour_scoring_log.txt',
    level=logging.DEBUG,
    format='%(asctime)s:%(levelname)s:%(message)s'
)

NL = '\n'

def get_fps(cap):
    """Return FPS if it looks valid, otherwise None.

    Some OpenCV backends on Windows may report bogus FPS values (e.g., 1000).
    In that case we return None so the user can enter the real FPS.
    """
    fps = cap.get(cv2.CAP_PROP_FPS)
    try:
        fps = float(fps)
    except Exception:
        return None
    if not fps or (isinstance(fps, float) and (np.isnan(fps) or fps <= 0)):
        return None
    # Heuristic guard: treat extremely high FPS as unreliable for typical behavioural videos.
    if fps > 240:
        return None
    return fps

def clamp(v, lo, hi):
    return max(lo, min(hi, v))


def open_video_capture(path: str) -> cv2.VideoCapture:
    """Open a VideoCapture with backends that tend to be more stable on Windows."""
    if sys.platform.startswith('win'):
        backend_candidates = []
        for name in ('CAP_FFMPEG', 'CAP_MSMF', 'CAP_DSHOW', 'CAP_ANY'):
            backend = getattr(cv2, name, None)
            if backend is not None:
                backend_candidates.append(backend)
        seen = set()
        for backend in backend_candidates:
            if backend in seen:
                continue
            seen.add(backend)
            cap = cv2.VideoCapture(path, backend)
            if cap.isOpened():
                return cap
            cap.release()
    return cv2.VideoCapture(path)


def rgb_numpy_to_qimage(rgb_image: np.ndarray) -> QImage:
    """Create a deep-copied QImage from an RGB (or RGBA/greyscale) NumPy array."""
    if rgb_image is None:
        return QImage()
    arr = np.ascontiguousarray(rgb_image)
    if arr.ndim == 2:
        arr = cv2.cvtColor(arr, cv2.COLOR_GRAY2RGB)
        arr = np.ascontiguousarray(arr)
    h, w = arr.shape[:2]
    ch = arr.shape[2] if arr.ndim == 3 else 1
    bytes_per_line = int(arr.strides[0])
    if ch == 3:
        return QImage(arr.data, w, h, bytes_per_line, QImage.Format_RGB888).copy()
    if ch == 4:
        fmt = getattr(QImage, 'Format_RGBA8888', QImage.Format_ARGB32)
        return QImage(arr.data, w, h, bytes_per_line, fmt).copy()
    arr = arr[:, :, :3]
    arr = np.ascontiguousarray(arr)
    h, w = arr.shape[:2]
    return QImage(arr.data, w, h, int(arr.strides[0]), QImage.Format_RGB888).copy()

class VideoThread(QThread):
    change_pixmap_signal = pyqtSignal(QImage)
    update_slider_signal = pyqtSignal(int)
    video_ended_signal = pyqtSignal()
    seek_signal = pyqtSignal(int)

    def __init__(self, video_path, playback_speed, fps, start_frame=0, annotations=None):
        super().__init__()
        self.video_path = video_path
        self.playback_speed = float(playback_speed)
        self.fps = float(fps) if fps else 30.0
        self.cap = open_video_capture(self.video_path)
        self.total_frames = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
        self._run_flag = True
        self.mutex = QMutex()
        self.annotations = annotations or []
        self.base_frame = int(start_frame)
        self.current_frame = int(start_frame)
        if self.cap.isOpened():
            self.cap.set(cv2.CAP_PROP_POS_FRAMES, self.base_frame)
        else:
            logging.error("Error opening video stream or file")
            self._run_flag = False
        self.start_wall = time.perf_counter()
        self.seek_signal.connect(self.seek_position)

    @pyqtSlot(int)
    def seek_position(self, position):
        with QMutexLocker(self.mutex):
            position = int(position)
            if 0 <= position < self.total_frames:
                self.cap.set(cv2.CAP_PROP_POS_FRAMES, position)
                self.base_frame = position
                self.current_frame = position
                self.start_wall = time.perf_counter()
                logging.info(f"Seeked to frame: {position}")
            else:
                logging.warning(f"Seek position {position} out of range.")

    def set_speed(self, new_speed: float):
        with QMutexLocker(self.mutex):
            self.playback_speed = max(0.1, min(8.0, float(new_speed)))
            self.base_frame = self.current_frame
            self.start_wall = time.perf_counter()

    def _sleep_until_next_frame(self):
        """Sleep until it's time to display the next frame.

        This is intentionally written to be responsive to stop/seek/speed changes:
        we sleep in small chunks and re-check the timing each time.
        """
        while True:
            with QMutexLocker(self.mutex):
                if not self._run_flag:
                    return
                spd = max(0.1, min(8.0, float(self.playback_speed)))
                fps = float(self.fps) if self.fps and float(self.fps) > 0 else 30.0
                # Target time (seconds since start_wall) for the next frame.
                next_index_rel = (self.current_frame + 1 - self.base_frame) / (fps * spd)
                now_rel = time.perf_counter() - self.start_wall
                dt = next_index_rel - now_rel

            if dt <= 0:
                return

            # Sleep in chunks so we stay responsive. 50 ms max keeps UI responsive and
            # avoids over-sleeping on Windows timer granularity.
            if dt >= 0.05:
                QThread.msleep(50)
            elif dt >= 0.002:
                QThread.msleep(max(1, int(dt * 1000)))
            else:
                # Sub-millisecond waits: use micro-sleep (best-effort; Windows may round).
                QThread.usleep(max(200, int(dt * 1_000_000)))

    def _catch_up_skip(self):

        with QMutexLocker(self.mutex):
            spd = max(0.1, min(8.0, self.playback_speed))
            now_rel = time.perf_counter() - self.start_wall
            target_index = int(now_rel * self.fps * spd) + self.base_frame
            target_index = min(target_index, self.total_frames - 1)
            delta = target_index - self.current_frame - 1
        if delta > 0:
            grabs = min(delta, 60)
            for _ in range(grabs):
                with QMutexLocker(self.mutex):
                    self.cap.grab()
            with QMutexLocker(self.mutex):
                self.current_frame += grabs

    def run(self):
        while self._run_flag:
            self._sleep_until_next_frame()
            self._catch_up_skip()
            with QMutexLocker(self.mutex):
                if not self._run_flag:
                    break
                ret, cv_img = self.cap.read()
            if not ret:
                self.video_ended_signal.emit()
                break
            with QMutexLocker(self.mutex):
                self.current_frame += 1
                frame_index = self.current_frame
                rgb_image = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
                frame_time = frame_index / self.fps if self.fps else 0.0
                for annotation in self.annotations:
                    if annotation['start_time'] <= frame_time <= annotation['end_time']:
                        behaviour = annotation['behaviour']
                        color = annotation.get('color', (255, 0, 0))
                        cv2.putText(rgb_image, behaviour, (50, 50), cv2.FONT_HERSHEY_SIMPLEX,
                                    1.5, color, 3, cv2.LINE_AA)


                qt_image = rgb_numpy_to_qimage(rgb_image)
            self.change_pixmap_signal.emit(qt_image)
            self.update_slider_signal.emit(frame_index)
        with QMutexLocker(self.mutex):
            self.cap.release()

    def stop(self):
        with QMutexLocker(self.mutex):
            self._run_flag = False
        self.wait()

class VideoWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.image = None

    def paintEvent(self, event):
        painter = QPainter(self)
        if self.image:
            rect = self.rect()
            img_rect = QRectF(0, 0, self.image.width(), self.image.height())
            target_rect = QRectF(0, 0, rect.width(), rect.height())
            painter.drawImage(target_rect, self.image, img_rect)
        else:
            painter.fillRect(self.rect(), Qt.black)

    def update_image(self, qt_image):
        self.image = qt_image
        self.update()

class TimelineWidget(QWidget):
    def __init__(self, events, total_duration):
        super().__init__()
        self.events = events or []
        self.total_duration = total_duration or 1.0
        self.current_time = 0.0
        self.setFixedHeight(22)
        self.setSizePolicy(QSizePolicy.Expanding, QSzPolicy.Fixed)

    def update_events(self, events):
        self.events = events or []
        self.update()

    def set_current_time(self, t):
        self.current_time = clamp(float(t), 0.0, self.total_duration if self.total_duration else 0.0)
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        rect = self.rect()
        width = rect.width()
        height = self.rect().height()
        painter.fillRect(rect, QColor("#2e2e2e"))
        behaviour_colors = {}
        palette = [
            QColor("#ff9999"), QColor("#99ff99"), QColor("#9999ff"),
            QColor("#ffff99"), QColor("#99ffff"), QColor("#ff99ff"),
            QColor("#ffcc99"), QColor("#ccff99"), QColor("#99ffcc"),
            QColor("#99ccff")
        ]
        phases = []
        for ev in self.events:
            ph = ev.get('phase', 'Default Phase')
            if ph not in phases:
                phases.append(ph)
        for idx, phase in enumerate(phases):
            behaviour_colors[phase] = palette[idx % len(palette)]
        for ev in self.events:
            phase = ev.get('phase', 'Default Phase')
            start_px = (ev['start_time'] / self.total_duration) * width if self.total_duration else 0
            end_px = (ev['end_time'] / self.total_duration) * width if self.total_duration else 0
            seg_h = max(4, height // 3)
            y_top = (height - seg_h) // 2
            painter.fillRect(QRectF(start_px, y_top, max(1.0, end_px - start_px), seg_h),
                             behaviour_colors.get(phase, QColor("#ffffff")))
        pos_x = 0 if not self.total_duration else (self.current_time / self.total_duration) * width
        pen = QPen(QColor('#ffffff'))
        pen.setWidth(2)
        painter.setPen(pen)
        painter.drawLine(int(pos_x), 0, int(pos_x), height)
        painter.end()

class HistoryPanel(QListWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('History Panel')
        self.setGeometry(50, 50, 300, 500)
        self.setStyleSheet('background-color: #2e2e2e; color: #ffffff;')

    def add_action(self, action_text):
        self.addItem(action_text)
        self.scrollToBottom()

class Configuration:
    def __init__(self):
        self.settings = QSettings('BehaviourScoringApp', 'Settings')

    def load_settings(self):
        try:
            if self.settings.contains('scoring_keys'):
                scoring_keys_json = self.settings.value('scoring_keys')
                raw = json.loads(scoring_keys_json)
                result = {}
                for k, v in raw.items():
                    try:
                        if v is None or v == -1:
                            continue
                        result[k] = int(v)
                    except Exception:
                        continue
                return result
            return {}
        except Exception as e:
            logging.error(f'Error loading settings: {e}')
            return {}

    def save_settings(self, scoring_keys):
        try:
            clean = {k: int(v) for k, v in scoring_keys.items() if isinstance(v, int)}
            scoring_keys_json = json.dumps(clean)
            self.settings.setValue('scoring_keys', scoring_keys_json)
        except Exception as e:
            logging.error(f'Error saving settings: {e}')

class AssignKeyDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle('Assign Keys and Behaviours')
        self.setMinimumSize(640, 520)
        self.behaviour_keys = self.parent().scoring_keys.copy()
        self.current_behaviour = None
        self.waiting_for_key = False
        outer = QVBoxLayout(self)
        self.instructions = QLabel("Enter behaviour name, click 'Add & Assign', then press a key...")
        self.instructions.setStyleSheet("color: white; font-size: 14px;")
        outer.addWidget(self.instructions)
        add_row = QHBoxLayout()
        self.behaviour_input = QLineEdit()
        self.behaviour_input.setPlaceholderText("Enter behaviour name")
        self.add_assign_btn = QPushButton('Add & Assign')
        self.add_assign_btn.setToolTip('Create a new behaviour and assign a key to it')
        self.add_assign_btn.clicked.connect(self.add_and_assign)
        for b in (self.add_assign_btn,):
            b.setStyleSheet("""
                QPushButton {
                    color: #ffffff;
                    background-color: #555555;
                    border: 2px solid #ffffff;
                    border-radius: 10px;
                    font-size: 14px;
                    padding: 5px 10px;
                }
                QPushButton:hover { background-color: #777777; }
            """)
        add_row.addWidget(self.behaviour_input, 1)
        add_row.addWidget(self.add_assign_btn, 0)
        outer.addLayout(add_row)
        self.assigned_table = QTableWidget()
        self.assigned_table.setColumnCount(2)
        self.assigned_table.setHorizontalHeaderLabels(['Behaviour', 'Key'])
        self.assigned_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.assigned_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.assigned_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.assigned_table.setStyleSheet("""
            QTableWidget {
                background-color: #2e2e2e;
                color: #ffffff;
                border: 1px solid #ffffff;
            }
            QHeaderView::section {
                background-color: #555555;
                color: #ffffff;
            }
        """)
        self.assigned_table.doubleClicked.connect(self.rename_selected)
        outer.addWidget(self.assigned_table)
        act_row = QHBoxLayout()
        self.assign_key_btn = QPushButton('Reassign Key')
        self.clear_key_btn = QPushButton('Clear Key')
        self.rename_btn = QPushButton('Rename')
        self.remove_btn = QPushButton('Remove')
        for b in (self.assign_key_btn, self.clear_key_btn, self.rename_btn, self.remove_btn):
            b.setStyleSheet("""
                QPushButton {
                    color: #ffffff;
                    background-color: #555555;
                    border: 2px solid #ffffff;
                    border-radius: 10px;
                    font-size: 13px;
                    padding: 4px 10px;
                }
                QPushButton:hover { background-color: #777777; }
            """)
        self.assign_key_btn.clicked.connect(self.reassign_selected)
        self.clear_key_btn.clicked.connect(self.clear_selected)
        self.rename_btn.clicked.connect(self.rename_selected)
        self.remove_btn.clicked.connect(self.remove_selected)
        act_row.addWidget(self.assign_key_btn)
        act_row.addWidget(self.clear_key_btn)
        act_row.addWidget(self.rename_btn)
        act_row.addWidget(self.remove_btn)
        act_row.addItem(QSpacerItem(10, 10, QSzPolicy.Expanding, QSzPolicy.Minimum))
        outer.addLayout(act_row)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        outer.addWidget(buttons)
        self.setFocusPolicy(Qt.StrongFocus)
        self.update_assigned_table()

    def _selected_behaviour(self):
        row = self.assigned_table.currentRow()
        if row < 0:
            return None
        return self.assigned_table.item(row, 0).text()

    def _confirm_replace_key(self, key_seq_str, existing_behaviour):
        return QMessageBox.question(
            self, 'Key Already Assigned',
            f"The key '{key_seq_str}' is already assigned to '{existing_behaviour}'.\n"
            f"Do you want to reassign it to the new behaviour?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        ) == QMessageBox.Yes

    def _find_behaviour_by_key(self, keycode):
        for b, k in self.behaviour_keys.items():
            if k == keycode:
                return b
        return None

    def add_and_assign(self):
        behaviour = self.behaviour_input.text().strip()
        if not behaviour:
            QMessageBox.warning(self, 'Input Error', 'Please enter a behaviour name.')
            return
        if behaviour in self.behaviour_keys:
            QMessageBox.warning(self, 'Duplicate Name', 'A behaviour with this name already exists.')
            return
        self.current_behaviour = behaviour
        self.waiting_for_key = True
        self.instructions.setText(f"Assigning for '{behaviour}'. Press any key now... (Esc to cancel)")
        self.add_assign_btn.setEnabled(False)
        self.grabKeyboard()
        self.setFocus()

    def reassign_selected(self):
        behaviour = self._selected_behaviour()
        if not behaviour:
            QMessageBox.information(self, 'No Selection', 'Please select a behaviour to reassign.')
            return
        self.current_behaviour = behaviour
        self.waiting_for_key = True
        self.instructions.setText(f"Reassign key for '{behaviour}'. Press any key now... (Esc to cancel)")
        self.grabKeyboard()
        self.setFocus()

    def clear_selected(self):
        behaviour = self._selected_behaviour()
        if not behaviour:
            QMessageBox.information(self, 'No Selection', 'Please select a behaviour to clear/remove.')
            return
        del self.behaviour_keys[behaviour]
        self.update_assigned_table()

    def rename_selected(self):
        behaviour = self._selected_behaviour()
        if not behaviour:
            return
        new_name, ok = QInputDialog.getText(self, 'Rename Behaviour', 'New name:', text=behaviour)
        if not ok:
            return
        new_name = new_name.strip()
        if not new_name:
            QMessageBox.warning(self, 'Invalid Name', 'Name cannot be empty.')
            return
        if new_name != behaviour and new_name in self.behaviour_keys:
            QMessageBox.warning(self, 'Duplicate Name', 'A behaviour with this name already exists.')
            return
        keycode = self.behaviour_keys[behaviour]
        del self.behaviour_keys[behaviour]
        self.behaviour_keys[new_name] = keycode
        self.update_assigned_table()

    def remove_selected(self):
        behaviour = self._selected_behaviour()
        if not behaviour:
            QMessageBox.information(self, 'No Selection', 'Please select a behaviour to remove.')
            return
        del self.behaviour_keys[behaviour]
        self.update_assigned_table()

    def keyPressEvent(self, event):
        if self.waiting_for_key:
            if event.key() == Qt.Key_Escape:
                self.waiting_for_key = False
                self.current_behaviour = None
                self.releaseKeyboard()
                self.add_assign_btn.setEnabled(True)
                self.instructions.setText("Key assignment cancelled. Enter behaviour and click 'Add & Assign'.")
                return
            key = event.key()
            modifiers = int(event.modifiers())
            combined_key = key + (modifiers << 16)
            key_seq_str = QKeySequence(combined_key).toString()
            existing = self._find_behaviour_by_key(combined_key)
            if existing and existing != self.current_behaviour:
                if not self._confirm_replace_key(key_seq_str, existing):
                    return
                del self.behaviour_keys[existing]
            if self.current_behaviour:
                self.behaviour_keys[self.current_behaviour] = combined_key
            self.current_behaviour = None
            self.waiting_for_key = False
            self.releaseKeyboard()
            self.add_assign_btn.setEnabled(True)
            self.behaviour_input.clear()
            self.instructions.setText("Assigned. Enter another behaviour or click Done.")
            self.update_assigned_table()
            return
        super().keyPressEvent(event)

    def closeEvent(self, event):
        try:
            if self.waiting_for_key:
                self.releaseKeyboard()
        except Exception:
            pass
        super().closeEvent(event)

    def update_assigned_table(self):
        self.assigned_table.setRowCount(0)
        for behaviour, key in self.behaviour_keys.items():
            row_position = self.assigned_table.rowCount()
            self.assigned_table.insertRow(row_position)
            self.assigned_table.setItem(row_position, 0, QTableWidgetItem(behaviour))
            self.assigned_table.setItem(row_position, 1, QTableWidgetItem(QKeySequence(key).toString()))

class TimeSpentDialog(QDialog):
    def __init__(self, parent, rows, total_seconds):
        super().__init__(parent)
        self.setWindowTitle('Time Spent (seconds)')
        self.setMinimumSize(480, 360)
        layout = QVBoxLayout(self)
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(['Behaviour', 'Seconds', 'Count', 'Percent'])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setStyleSheet("""
            QTableWidget {
                background-color: #2e2e2e;
                color: #ffffff;
                border: 1px solid #ffffff;
            }
            QHeaderView::section {
                background-color: #555555;
                color: #ffffff;
            }
        """)
        table.setRowCount(len(rows) + 1)
        for i, r in enumerate(rows):
            table.setItem(i, 0, QTableWidgetItem(r['behaviour']))
            table.setItem(i, 1, QTableWidgetItem(f"{r['seconds']:.3f}"))
            table.setItem(i, 2, QTableWidgetItem(str(r['count'])))
            pct = (r['seconds'] / total_seconds * 100.0) if total_seconds > 0 else 0.0
            table.setItem(i, 3, QTableWidgetItem(f"{pct:.1f}%"))
        total_count = sum(r['count'] for r in rows)
        table.setItem(len(rows), 0, QTableWidgetItem('Total'))
        table.setItem(len(rows), 1, QTableWidgetItem(f"{total_seconds:.3f}"))
        table.setItem(len(rows), 2, QTableWidgetItem(str(total_count)))
        table.setItem(len(rows), 3, QTableWidgetItem("100.0%" if total_seconds > 0 else "0.0%"))
        layout.addWidget(table)
        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        buttons.rejected.connect(self.reject)
        buttons.accepted.connect(self.accept)
        layout.addWidget(buttons)

class BehaviourScoringApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Rodent Manual Scorer')
        self.setGeometry(100, 100, 1400, 800)
        self.video_path = ''
        self.playback_speed = 1.0
        self.fps = None
        self.scoring_keys = {}
        self.time_spent = {}
        self.frequency_counts = {}
        self.scoring = False
        self.key_states = {}
        self.last_key_times = {}
        self.behaviour_events = []
        self.paused_frame = None
        self.undo_stack = []
        self.redo_stack = []
        self.current_phase = "Default Phase"
        self.phase_colors = {"Default Phase": QColor("#ffffff")}
        self.annotations = []
        self.is_fullscreen = False
        self.total_duration = 0.0
        palette = self.palette()
        palette.setColor(QPalette.Window, QColor("#1e1e1e"))
        self.setPalette(palette)
        self.history_panel = HistoryPanel()
        self.history_panel.hide()
        self.config = Configuration()
        self.scoring_keys = self.config.load_settings()
        self.initUI()
        self.installEventFilter(self)
        if QApplication.instance() is not None:
            QApplication.instance().installEventFilter(self)
        self.autosave_timer = QTimer()
        self.autosave_timer.timeout.connect(self.autosave_session)
        self.autosave_timer.start(300000)
        self.video_thread = None

    def initUI(self):
        self.main_widget = QWidget()
        self.setCentralWidget(self.main_widget)
        self.main_layout = QHBoxLayout(self.main_widget)
        self.left_layout = QVBoxLayout()
        self.right_layout = QVBoxLayout()
        self.main_layout.addLayout(self.left_layout, 3)
        self.main_layout.addLayout(self.right_layout, 1)
        self.create_menu()
        self.create_video_display()
        self.create_controls()
        self.create_behaviour_controls()
        self.create_frequency_controls()
        self.create_phase_manager()
        self.scoring_timer = QTimer()
        self.scoring_timer.timeout.connect(self.update_timers)
        self.setAcceptDrops(True)

    def create_menu(self):
        menubar = self.menuBar()
        file_menu = menubar.addMenu('File')
        load_video_action = QAction('Load Video', self)
        load_video_action.setShortcut('Ctrl+O')
        load_video_action.triggered.connect(self.upload_video)
        file_menu.addAction(load_video_action)
        load_csv_action = QAction('Load Scoring CSV', self)
        load_csv_action.setShortcut('Ctrl+L')
        load_csv_action.triggered.connect(self.load_scoring_csv)
        file_menu.addAction(load_csv_action)
        save_csv_action = QAction('Save Scoring CSV', self)
        save_csv_action.setShortcut('Ctrl+S')
        save_csv_action.triggered.connect(self.save_scoring_csv)
        file_menu.addAction(save_csv_action)
        export_excel_action = QAction('Export Excel', self)
        export_excel_action.setShortcut('Ctrl+E')
        export_excel_action.triggered.connect(self.export_excel)
        file_menu.addAction(export_excel_action)
        reset_action = QAction('New Session', self)
        reset_action.setShortcut('Ctrl+N')
        reset_action.triggered.connect(self.new_session)
        file_menu.addAction(reset_action)
        exit_action = QAction('Exit', self)
        exit_action.setShortcut('Ctrl+Q')
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        edit_menu = menubar.addMenu('Edit')
        undo_action = QAction('Undo Last (Ctrl+Z)', self)
        undo_action.setShortcut(QKeySequence.Undo)
        undo_action.triggered.connect(self.undo_last_event)
        edit_menu.addAction(undo_action)
        redo_action = QAction('Redo (Ctrl+Y)', self)
        redo_action.setShortcut('Ctrl+Y')
        redo_action.triggered.connect(self.redo_last_event)
        edit_menu.addAction(redo_action)

        view_menu = menubar.addMenu('View')
        history_action = QAction('Show History', self)
        history_action.triggered.connect(self.toggle_history_panel)
        view_menu.addAction(history_action)
        time_spent_action = QAction('Show Time Spent', self)
        time_spent_action.setShortcut('Ctrl+T')
        time_spent_action.triggered.connect(self.show_time_spent)
        view_menu.addAction(time_spent_action)
        fullscreen_action = QAction('Toggle Fullscreen', self)
        fullscreen_action.setShortcut('F11')
        fullscreen_action.triggered.connect(self.toggle_fullscreen)
        view_menu.addAction(fullscreen_action)

    def toggle_history_panel(self):
        if self.history_panel.isVisible():
            self.history_panel.hide()
        else:
            self.history_panel.show()

    def create_video_display(self):
        self.video_widget = VideoWidget()
        self.video_widget.setMinimumSize(1000, 600)
        self.video_widget.setSizePolicy(QSizePolicy.Expanding, QSzPolicy.Expanding)
        self.video_widget.setStyleSheet("""
            QWidget {
                background-color: #000000;
                border: 1px solid #ffffff;
                border-radius: 15px;
            }
        """)
        self.left_layout.addWidget(self.video_widget)

    def create_controls(self):
        self.controls_layout = QHBoxLayout()
        self.left_layout.addLayout(self.controls_layout)
        self.play_button = self.create_button('Play', self.play_video)
        self.controls_layout.addWidget(self.play_button)
        self.pause_button = self.create_button('Pause', self.pause_video)
        self.controls_layout.addWidget(self.pause_button)
        self.stop_button = self.create_button('Stop', self.stop_video)
        self.controls_layout.addWidget(self.stop_button)
        self.fullscreen_button = self.create_button('Fullscreen', self.toggle_fullscreen)
        self.controls_layout.addWidget(self.fullscreen_button)
        self.controls_layout.addStretch()
        self.speed_dropdown = QComboBox()
        speeds = ['0.1x', '0.25x', '0.5x', '0.75x', '1x', '1.25x', '1.5x', '2x', '3x', '4x', '6x', '8x']
        self.speed_dropdown.addItems(speeds)
        self.speed_dropdown.setCurrentText('1x')
        self.speed_dropdown.currentTextChanged.connect(self.change_playback_speed)
        self.speed_dropdown.setToolTip('Playback Speed')
        self.speed_dropdown.setStyleSheet("""
            QComboBox {
                background-color: #333333;
                color: #ffffff;
                border: 1px solid #ffffff;
                border-radius: 5px;
                padding: 5px;
                min-width: 96px;
            }
            QComboBox:hover {
                background-color: #555555;
            }
        """)
        self.controls_layout.addWidget(self.speed_dropdown)
        self.video_slider = QSlider(Qt.Horizontal)
        self.video_slider.setStyleSheet("""
            QSlider::groove:horizontal {
                border: 1px solid #999999;
                height: 8px;
                background: #333333;
                margin: 2px 0;
            }
            QSlider::handle:horizontal {
                background: #ffffff;
                border: 1px solid #ffffff;
                width: 18px;
                margin: -2px 0;
                border-radius: 3px;
            }
        """)
        self.video_slider.sliderMoved.connect(self.set_video_position)
        self.left_layout.addWidget(self.video_slider)
        self.timeline_widget = TimelineWidget([], 1.0)
        self.left_layout.addWidget(self.timeline_widget)
        self.timestamp_label = QLabel('00:00:00.000 / 00:00:00.000')
        self.timestamp_label.setAlignment(Qt.AlignCenter)
        self.timestamp_label.setStyleSheet('color: #ffffff;')
        self.left_layout.addWidget(self.timestamp_label)

    def toggle_fullscreen(self):
        try:
            if not self.is_fullscreen:
                self.showFullScreen()
                self.is_fullscreen = True
            else:
                self.showNormal()
                self.is_fullscreen = False
        except Exception as e:
            logging.error(f'Error toggling fullscreen: {e}')

    def create_button(self, text, function):
        button = QPushButton(text)
        button.setMinimumSize(100, 40)
        button.setSizePolicy(QSizePolicy.Fixed, QSzPolicy.Fixed)
        button.setStyleSheet("""
            QPushButton {
                color: #ffffff;
                background-color: transparent;
                border: 2px solid #ffffff;
                border-radius: 20px;
                font-size: 14px;
                padding: 5px 14px;
            }
            QPushButton:hover {
                background-color: #333333;
            }
        """)
        button.clicked.connect(function)
        return button

    def create_behaviour_controls(self):
        self.behaviour_label = QLabel('Behaviours:')
        self.behaviour_label.setStyleSheet('color: #ffffff; font-size: 18px;')
        self.right_layout.addWidget(self.behaviour_label)
        self.behaviour_list = QListWidget()
        self.behaviour_list.setStyleSheet('background-color: #2e2e2e; color: #ffffff;')
        self.right_layout.addWidget(self.behaviour_list)
        self.assign_keys_button = QPushButton('Assign Keys')
        self.assign_keys_button.setStyleSheet("""
            QPushButton {
                color: #ffffff;
                background-color: #555555;
                border: 2px solid #ffffff;
                border-radius: 10px;
                font-size: 14px;
                padding: 5px;
            }
            QPushButton:hover {
                background-color: #777777;
            }
        """)
        self.assign_keys_button.clicked.connect(self.open_assign_dialog)
        self.right_layout.addWidget(self.assign_keys_button)
        self.feedback_label = QLabel('')
        self.feedback_label.setStyleSheet('color: #ffffff; font-size: 14px;')
        self.right_layout.addWidget(self.feedback_label)
        self.behaviour_labels = {}

    def create_frequency_controls(self):
        self.frequency_groupbox = QGroupBox("Frequency Display Options")
        self.frequency_layout = QVBoxLayout()
        self.frequency_groupbox.setLayout(self.frequency_layout)
        self.global_frequency_checkbox = QCheckBox("Enable Frequency Counts")
        self.global_frequency_checkbox.setChecked(True)
        self.global_frequency_checkbox.stateChanged.connect(self.toggle_global_frequency)
        self.frequency_layout.addWidget(self.global_frequency_checkbox)
        self.right_layout.addWidget(self.frequency_groupbox)

    def create_phase_manager(self):
        self.phase_manager_layout = QVBoxLayout()
        self.phase_button = QPushButton('Start New Phase')
        self.phase_button.setStyleSheet("""
            QPushButton {
                color: #ffffff;
                background-color: #555555;
                border: 2px solid #ffffff;
                border-radius: 10px;
                font-size: 14px;
                padding: 5px;
            }
            QPushButton:hover {
                background-color: #777777;
            }
        """)
        self.phase_button.clicked.connect(self.start_new_phase)
        self.phase_manager_layout.addWidget(self.phase_button)
        self.phase_color_button = QPushButton('Select Phase Colour')
        self.phase_color_button.setStyleSheet("""
            QPushButton {
                color: #ffffff;
                background-color: #555555;
                border: 2px solid #ffffff;
                border-radius: 10px;
                font-size: 14px;
                padding: 5px;
            }
            QPushButton:hover {
                background-color: #777777;
            }
        """)
        self.phase_color_button.setVisible(False)
        self.phase_color_button.clicked.connect(self.select_phase_color)
        self.phase_manager_layout.addWidget(self.phase_color_button)
        self.selected_color = QColor("#ffffff")
        self.right_layout.addLayout(self.phase_manager_layout)

    def upload_video(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(
            self, 'Select Video File', '',
            'Video Files (*.avi *.mp4 *.mov *.mkv);;All Files (*)', options=options
        )
        if file_name:
            self.load_video(file_name)

    def load_video(self, file_name):
        try:
            if self.video_thread and self.video_thread.isRunning():
                self.video_thread.stop()
            self.scoring_timer.stop()
            self.scoring = False
            self.time_spent = {}
            self.frequency_counts = {}
            self.behaviour_events = []
            self.undo_stack = []
            self.redo_stack = []
            self.annotations = []
            self.paused_frame = None
            self.video_path = file_name
            self.cap = open_video_capture(self.video_path)
            if not self.cap.isOpened():
                QMessageBox.critical(self, 'Error', 'Failed to open the video file.')
                logging.error('Failed to open the video file.')
                return
            total_frames = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
            self.video_slider.setMaximum(total_frames if total_frames > 0 else 0)
            fps = get_fps(self.cap)
            if fps is None:
                fps_input, ok = QInputDialog.getDouble(
                    self, 'FPS Required', 'The video file does not report FPS. Please enter the FPS:',
                    decimals=2, min=1.0, value=30.0
                )
                if ok and fps_input > 0:
                    self.fps = float(fps_input)
                else:
                    QMessageBox.warning(self, 'Error', 'Valid FPS is required. Cannot proceed without FPS.')
                    logging.error('Invalid FPS input.')
                    self.cap.release()
                    return
            else:
                self.fps = float(fps)
            total_seconds = (total_frames / self.fps) if self.fps else 0.0
            self.total_duration = total_seconds
            self.timestamp_label.setText(f'00:00:00.000 / {self.format_time(total_seconds)}')
            self.cap.release()
            logging.info(f'Video loaded: {self.video_path}')
            self.feedback_label.setText('Video loaded. Assign keys and start scoring.')
            self.paused_frame = None
            self.video_slider.setValue(0)
            self.update_timestamp(0, total_seconds)
            self.video_widget.image = None
            self.video_widget.update()
            self.timeline_widget.total_duration = total_seconds
        except Exception as e:
            QMessageBox.warning(self, 'Error', f'Failed to load video: {e}')
            logging.error(f'Failed to load video: {e}')

    def play_video(self):
        try:
            if not self.video_path:
                QMessageBox.warning(self, 'No Video', 'Please load a video first.')
                return
            start_frame = self.paused_frame if self.paused_frame is not None else 0
            self.video_thread = VideoThread(self.video_path, self.playback_speed, self.fps, start_frame, self.annotations)
            self.video_thread.change_pixmap_signal.connect(self.update_image)
            self.video_thread.update_slider_signal.connect(self.update_slider)
            self.video_thread.video_ended_signal.connect(self.on_video_end)
            self.video_thread.start()
            self.scoring_timer.start(50)
            self.scoring = True
            self.feedback_label.setText('Scoring in progress...')
            logging.info('Video playback started.')
            self.paused_frame = None
        except Exception as e:
            QMessageBox.warning(self, 'Error', f'Failed to start video: {e}')
            logging.error(f'Failed to start video: {e}')

    def pause_video(self):
        try:
            if self.video_thread:
                if self.video_thread.isRunning():
                    self.paused_frame = self.video_thread.current_frame
                    self.video_thread.stop()
                    self.scoring_timer.stop()
                    self.scoring = False
                    self.feedback_label.setText('Paused.')
                    logging.info('Video paused.')
                else:
                    self.play_video()
            else:
                self.play_video()
        except Exception as e:
            logging.error(f'Error pausing video: {e}')

    def stop_video(self):
        try:
            if self.video_thread:
                self.video_thread.stop()
                self.video_thread = None
            self.scoring_timer.stop()
            self.scoring = False
            self.paused_frame = 0
            self.video_slider.setValue(0)
            self.update_timestamp(0, self.total_duration)
            self.update_video_display(0)
            self.feedback_label.setText('Stopped.')
            logging.info('Video stopped.')
        except Exception as e:
            logging.error(f'Error stopping video: {e}')

    def on_video_end(self):
        try:
            self.scoring_timer.stop()
            self.scoring = False
            self.feedback_label.setText('Video ended.')
            logging.info('Video ended.')
        except Exception as e:
            logging.error(f'Error handling video end: {e}')

    def update_image(self, qt_image):
        try:
            self.video_widget.update_image(qt_image)
        except Exception as e:
            logging.error(f'Error updating image: {e}')

    def change_playback_speed(self, *_):
        try:
            speed_text = self.speed_dropdown.currentText()
            self.playback_speed = float(speed_text.replace('x', ''))
            self.playback_speed = max(0.1, min(8.0, self.playback_speed))
            if self.video_thread and self.video_thread.isRunning():
                self.video_thread.set_speed(self.playback_speed)
            logging.info(f'Playback speed set to {self.playback_speed}x.')
        except Exception as e:
            logging.error(f'Error changing playback speed: {e}')

    def set_video_position(self, position):
        try:
            if not self.fps:
                return
            position = int(position)
            if self.video_thread and self.video_thread.isRunning():
                if 0 <= position <= self.video_slider.maximum():
                    self.video_thread.seek_signal.emit(position)
                    self.paused_frame = position
                    self.video_slider.setValue(position)
                    self.update_timestamp(position / self.fps)
                    self.timeline_widget.set_current_time(position / self.fps)
                    # Do not open a second VideoCapture while the playback thread is running (Windows can crash);
                    # the thread will emit the next frame after seeking.
                else:
                    QMessageBox.warning(self, 'Invalid Position', 'The selected position is out of range.')
                    logging.warning(f'Set video position out of range: {position}')
            else:
                if 0 <= position <= self.video_slider.maximum():
                    self.paused_frame = position
                    self.video_slider.setValue(position)
                    self.update_timestamp(position / self.fps)
                    self.timeline_widget.set_current_time(position / self.fps)
                    self.update_video_display(position)
                else:
                    QMessageBox.warning(self, 'Invalid Position', 'The selected position is out of range.')
                    logging.warning(f'Set video position out of range: {position}')
        except Exception as e:
            logging.error(f'Error setting video position: {e}')

    def update_video_display(self, frame_number=None):
        try:
            if not self.video_path:
                return
            if frame_number is None:
                frame_number = self.paused_frame or 0
            cap = open_video_capture(self.video_path)
            try:
                cap.set(cv2.CAP_PROP_POS_FRAMES, int(frame_number))
                ret, cv_img = cap.read()
                if ret:
                    rgb_image = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
                    frame_time = frame_number / self.fps if self.fps else 0.0
                    for annotation in self.annotations:
                        if annotation['start_time'] <= frame_time <= annotation['end_time']:
                            behaviour = annotation['behaviour']
                            color = annotation.get('color', (255, 0, 0))
                            cv2.putText(rgb_image, behaviour, (50, 50), cv2.FONT_HERSHEY_SIMPLEX,
                                        1.5, color, 3, cv2.LINE_AA)


                    qt_image = rgb_numpy_to_qimage(rgb_image)
                    self.video_widget.update_image(qt_image)
                self.video_widget.update()
                if self.fps:
                    self.timeline_widget.set_current_time((frame_number or 0) / self.fps)
            finally:
                cap.release()
        except Exception as e:
            logging.error(f'Error updating video display: {e}')

    def update_slider(self, position):
        try:
            self.video_slider.blockSignals(True)
            self.video_slider.setValue(int(position))
            self.video_slider.blockSignals(False)
            current_time = (int(position) / self.fps) if self.fps else 0.0
            total_time = (self.video_slider.maximum() / self.fps) if self.fps else 0.0
            self.update_timestamp(current_time, total_time)
            self.timeline_widget.set_current_time(current_time)
        except Exception as e:
            logging.error(f'Error updating slider: {e}')

    def update_timestamp(self, current_time, total_time=None):
        try:
            if total_time is None:
                total_time = (self.video_slider.maximum() / self.fps) if self.fps else 0.0
            self.timestamp_label.setText(f'{self.format_time(current_time)} / {self.format_time(total_time)}')
        except Exception as e:
            logging.error(f'Error updating timestamp: {e}')

    def _accurate_video_time(self):
        try:
            if self.video_thread and self.video_thread.isRunning() and self.fps:
                return self.video_thread.current_frame / self.fps
            return (self.video_slider.value() / self.fps) if self.fps else 0.0
        except Exception:
            return (self.video_slider.value() / self.fps) if self.fps else 0.0

    def format_time(self, seconds):
        try:
            seconds = float(seconds)
            millis = int((seconds - int(seconds)) * 1000)
            seconds = int(seconds)
            mins, secs = divmod(seconds, 60)
            hours, mins = divmod(mins, 60)
            return f'{hours:02}:{mins:02}:{secs:02}.{millis:03}'
        except Exception as e:
            logging.error(f'Error formatting time: {e}')
            return '00:00:00.000'

    def open_assign_dialog(self):
        try:
            dialog = AssignKeyDialog(self)
            if dialog.exec_():
                self.scoring_keys = dialog.behaviour_keys
                self.update_behaviour_labels()
                self.update_frequency_controls()
                self.config.save_settings(self.scoring_keys)
                logging.info('Scoring keys updated.')
        except Exception as e:
            logging.error(f'Error opening assign dialog: {e}')

    def update_behaviour_labels(self):
        try:
            for label in getattr(self, 'behaviour_labels', {}).values():
                self.right_layout.removeWidget(label)
                label.deleteLater()
            self.behaviour_labels = {}
            for behaviour, key in self.scoring_keys.items():
                time_spent = self.time_spent.get(behaviour, 0.0)
                freq = self.frequency_counts.get(behaviour, 0)
                label = QLabel(
                    f'{behaviour}: {QKeySequence(key).toString()} | Time: {self.format_time(time_spent)} | Count: {freq}'
                )
                label.setStyleSheet('color: #ffffff; font-size: 14px;')
                self.right_layout.addWidget(label)
                self.behaviour_labels[behaviour] = label
        except Exception as e:
            logging.error(f'Error updating behaviour labels: {e}')

    def update_frequency_controls(self):
        try:
            for i in reversed(range(self.frequency_layout.count())):
                widget = self.frequency_layout.itemAt(i).widget()
                if widget and widget != self.global_frequency_checkbox:
                    self.frequency_layout.removeWidget(widget)
                    widget.deleteLater()
            self.frequency_checkboxes = {}
            for behaviour in self.scoring_keys.keys():
                checkbox = QCheckBox(behaviour)
                checkbox.setChecked(True)
                self.frequency_layout.addWidget(checkbox)
                self.frequency_checkboxes[behaviour] = checkbox
        except Exception as e:
            logging.error(f'Error updating frequency controls: {e}')

    def toggle_global_frequency(self, state):
        try:
            enabled = state == Qt.Checked
            for checkbox in getattr(self, 'frequency_checkboxes', {}).values():
                checkbox.setChecked(enabled)
        except Exception as e:
            logging.error(f'Error toggling global frequency: {e}')

    def start_new_phase(self):
        try:
            text, ok = QInputDialog.getText(self, 'New Phase', 'Enter the name of the new phase:')
            if ok and text:
                phase_name = text.strip()
                if phase_name in self.phase_colors:
                    QMessageBox.warning(self, 'Duplicate Phase', 'A phase with this name already exists.')
                    return
                color = QColorDialog.getColor(self.phase_colors.get(self.current_phase, QColor("#ffffff")), self, 'Select Phase Colour')
                if color.isValid():
                    self.phase_colors[phase_name] = color
                    self.current_phase = phase_name
                    self.history_panel.add_action(f'New phase started: {phase_name}')
                    logging.info(f'New phase started: {phase_name}')
                else:
                    QMessageBox.warning(self, 'Invalid Colour', 'Please select a valid colour.')
        except Exception as e:
            logging.error(f'Error starting new phase: {e}')

    def select_phase_color(self):
        try:
            color = QColorDialog.getColor(self.phase_colors.get(self.current_phase, QColor("#ffffff")), self, 'Select Phase Colour')
            if color.isValid():
                self.phase_colors[self.current_phase] = color
                logging.info(f'Phase colour updated for {self.current_phase}')
        except Exception as e:
            logging.error(f'Error selecting phase colour: {e}')

    def show_time_spent(self):
        try:
            if not self.time_spent:
                QMessageBox.information(self, 'No Data', 'No time has been recorded yet.')
                return
            total_seconds = sum(float(v) for v in self.time_spent.values())
            rows = []
            for behaviour, secs in sorted(self.time_spent.items(), key=lambda x: -x[1]):
                rows.append({
                    'behaviour': behaviour,
                    'seconds': float(secs),
                    'count': int(self.frequency_counts.get(behaviour, 0))
                })
            dlg = TimeSpentDialog(self, rows, total_seconds)
            dlg.exec_()
        except Exception as e:
            logging.error(f'Error showing time spent: {e}')
            QMessageBox.warning(self, 'Error', f'Failed to show time spent: {e}')

    def update_timers(self):
        try:
            _ = self._accurate_video_time()
        except Exception as e:
            logging.error(f'Error updating timers: {e}')

    def get_video_time(self):
        return self._accurate_video_time()

    def keyPressEvent(self, event):
        super().keyPressEvent(event)

    def keyReleaseEvent(self, event):
        super().keyReleaseEvent(event)

    def undo_last_event(self):
        try:
            if not self.behaviour_events:
                QMessageBox.information(self, 'Undo', 'No scored behaviours to undo.')
                return
            ev = self.behaviour_events.pop()
            self.redo_stack.append(ev)
            b = ev['behaviour']
            dur = float(ev.get('duration', 0.0))
            self.time_spent[b] = max(0.0, self.time_spent.get(b, 0.0) - dur)
            if b in self.frequency_counts and self.frequency_counts[b] > 0:
                self.frequency_counts[b] -= 1
            self.update_behaviour_label(b)
            for i in range(len(self.annotations) - 1, -1, -1):
                a = self.annotations[i]
                if (a.get('behaviour') == b and
                    abs(a.get('start_time', -1) - ev['start_time']) < 1e-3 and
                    abs(a.get('end_time', -1) - ev['end_time']) < 1e-3):
                    self.annotations.pop(i)
                    break
            self.timeline_widget.update_events(self.get_all_behaviour_events())
            if self.video_thread and self.video_thread.isRunning():
                self.video_thread.stop()
            self.scoring_timer.stop()
            self.scoring = False
            start_frame = int(round(ev['start_time'] * self.fps)) if self.fps else 0
            self.paused_frame = start_frame
            self.video_slider.setValue(start_frame)
            self.update_timestamp(ev['start_time'])
            self.timeline_widget.set_current_time(ev['start_time'])
            self.update_video_display(start_frame)
            msg = f"Undid '{b}'. Rewound to {self.format_time(ev['start_time'])}."
            self.feedback_label.setText(msg)
            self.history_panel.add_action(f'UNDO {b} at {self.format_time(ev["start_time"])}')
            logging.info(msg)
        except Exception as e:
            logging.error(f'Error during undo: {e}')
            QMessageBox.warning(self, 'Error', f'Failed to undo: {e}')

    def redo_last_event(self):
        try:
            if not self.redo_stack:
                QMessageBox.information(self, 'Redo', 'Nothing to redo.')
                return
            ev = self.redo_stack.pop()
            b = ev['behaviour']
            self.behaviour_events.append(ev)
            self.time_spent[b] = self.time_spent.get(b, 0.0) + float(ev.get('duration', 0.0))
            self.frequency_counts[b] = self.frequency_counts.get(b, 0) + 1
            self.update_behaviour_label(b)
            self.annotations.append({
                'behaviour': b,
                'start_time': ev['start_time'],
                'end_time': ev['end_time'],
                'color': self.phase_colors.get(ev.get('phase', 'Default Phase'), QColor("#ffffff")).getRgb()[:3]
            })
            self.timeline_widget.update_events(self.get_all_behaviour_events())
            self.history_panel.add_action(f'REDO {b} at {self.format_time(ev["start_time"])}')
            self.feedback_label.setText(f"Redid '{b}'.")
            logging.info(f"Redid '{b}'.")
        except Exception as e:
            logging.error(f'Error during redo: {e}')
            QMessageBox.warning(self, 'Error', f'Failed to redo: {e}')

    def eventFilter(self, source, event):
        if self.focusWidget() and isinstance(self.focusWidget(), QLineEdit):
            return False
        if QApplication.activeModalWidget():
            return False
        if event.type() == QEvent.KeyPress:
            if (event.modifiers() & Qt.ControlModifier) and event.key() == Qt.Key_Z:
                self.undo_last_event()
                return True
            if (event.modifiers() & Qt.ControlModifier) and event.key() == Qt.Key_Y:
                self.redo_last_event()
                return True
            if event.key() == Qt.Key_Space:
                if self.video_thread and self.video_thread.isRunning():
                    self.pause_video()
                else:
                    self.play_video()
                return True
            if event.key() == Qt.Key_P:
                if self.scoring:
                    self.start_new_phase()
                return True
            if self.scoring:
                if event.isAutoRepeat():
                    return True
                key = event.key()
                modifiers = int(event.modifiers())
                combined_key = key + (modifiers << 16)
                if combined_key in self.scoring_keys.values() and combined_key not in self.key_states:
                    self.key_states[combined_key] = True
                    self.last_key_times[combined_key] = self._accurate_video_time()
                    self.highlight_behaviour_label(combined_key, True)
            return True
        elif event.type() == QEvent.KeyRelease:
            if self.scoring:
                if event.isAutoRepeat():
                    return True
                key = event.key()
                modifiers = int(event.modifiers())
                combined_key = key + (modifiers << 16)
                if combined_key in self.scoring_keys.values() and combined_key in self.key_states:
                    start_time = self.last_key_times.pop(combined_key)
                    end_time = self._accurate_video_time()
                    duration = max(0.0, end_time - start_time)
                    behaviour = None
                    for b, assigned_key in self.scoring_keys.items():
                        if assigned_key == combined_key:
                            behaviour = b
                            event_record = {
                                'phase': self.current_phase,
                                'behaviour': behaviour,
                                'start_time': start_time,
                                'end_time': end_time,
                                'duration': duration
                            }
                            self.behaviour_events.append(event_record)
                            self.undo_stack.append(event_record)
                            self.redo_stack.clear()
                            self.time_spent[behaviour] = self.time_spent.get(behaviour, 0.0) + duration
                            self.frequency_counts[behaviour] = self.frequency_counts.get(behaviour, 0) + 1
                            self.update_behaviour_label(behaviour)
                            self.history_panel.add_action(
                                f'Added {behaviour}: {self.format_time(start_time)} to {self.format_time(end_time)}'
                            )
                            break
                    if behaviour:
                        self.highlight_behaviour_label(combined_key, False)
                        self.timeline_widget.update_events(self.get_all_behaviour_events())
                        self.annotations.append({
                            'behaviour': behaviour,
                            'start_time': start_time,
                            'end_time': end_time,
                            'color': self.phase_colors.get(self.current_phase, QColor("#ffffff")).getRgb()[:3]
                        })
                    self.key_states.pop(combined_key, None)
            return True
        return super().eventFilter(source, event)

    def highlight_behaviour_label(self, combined_key, highlight):
        try:
            for behaviour, assigned_key in self.scoring_keys.items():
                if assigned_key == combined_key and behaviour in self.behaviour_labels:
                    if highlight:
                        self.behaviour_labels[behaviour].setStyleSheet('color: #00ff00; font-size: 14px;')
                    else:
                        self.behaviour_labels[behaviour].setStyleSheet('color: #ffffff; font-size: 14px;')
        except Exception as e:
            logging.error(f'Error highlighting behaviour label: {e}')

    def update_behaviour_label(self, behaviour):
        try:
            if behaviour in self.behaviour_labels and behaviour in self.scoring_keys:
                time_spent = self.time_spent.get(behaviour, 0.0)
                freq = self.frequency_counts.get(behaviour, 0)
                self.behaviour_labels[behaviour].setText(
                    f'{behaviour}: {QKeySequence(self.scoring_keys[behaviour]).toString()} '
                    f'| Time: {self.format_time(time_spent)} | Count: {freq}'
                )
        except Exception as e:
            logging.error(f'Error updating behaviour label: {e}')

    def load_scoring_csv(self):
        try:
            options = QFileDialog.Options()
            file_name, _ = QFileDialog.getOpenFileName(self, 'Open Scoring CSV', '', 'CSV Files (*.csv);;All Files (*)', options=options)
            if file_name:
                self.behaviour_events = []
                self.time_spent = {}
                self.frequency_counts = {}
                with open(file_name, 'r', newline='', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        try:
                            phase = (row.get('Phase') or row.get('phase') or 'Default Phase').strip() or 'Default Phase'
                            behaviour = (row.get('Behaviour') or row.get('behaviour') or '').strip()
                            if not behaviour:
                                continue
                            start_raw = row.get('Start Time') or row.get('start_time') or row.get('Start') or ''
                            end_raw = row.get('End Time') or row.get('end_time') or row.get('End') or ''
                            dur_raw = row.get('Duration (s)') or row.get('duration') or row.get('Duration') or ''
                            start_time = self.parse_time_to_seconds(start_raw)
                            end_time = self.parse_time_to_seconds(end_raw)
                            try:
                                duration = float(dur_raw)
                            except Exception:
                                duration = max(0.0, end_time - start_time)
                            event = {
                                'phase': phase,
                                'behaviour': behaviour,
                                'start_time': float(start_time),
                                'end_time': float(end_time),
                                'duration': float(duration)
                            }
                            self.behaviour_events.append(event)
                            self.time_spent[behaviour] = self.time_spent.get(behaviour, 0.0) + event['duration']
                            self.frequency_counts[behaviour] = self.frequency_counts.get(behaviour, 0) + 1
                        except Exception as inner:
                            logging.warning(f'Skipping bad row in CSV: {inner} | row={row}')
                self.update_behaviour_labels()
                self.update_frequency_controls()
                self.timeline_widget.update_events(self.get_all_behaviour_events())
                QMessageBox.information(self, 'Loaded', 'Scoring CSV loaded successfully.')
        except Exception as e:
            QMessageBox.warning(self, 'Error', f'Failed to load CSV: {e}')
            logging.error(f'Failed to load scoring CSV: {e}')

    def save_scoring_csv(self):
        try:
            options = QFileDialog.Options()
            file_name, _ = QFileDialog.getSaveFileName(self, 'Save Scoring CSV', '', 'CSV Files (*.csv);;All Files (*)', options=options)
            if file_name:
                if not file_name.lower().endswith('.csv'):
                    file_name += '.csv'
                self.write_csv(file_name)
                QMessageBox.information(self, 'Saved', f'Scoring CSV saved to {file_name}')
        except Exception as e:
            QMessageBox.warning(self, 'Error', f'Failed to save CSV: {e}')
            logging.error(f'Failed to save scoring CSV: {e}')

    def write_csv(self, file_name):
        try:
            with open(file_name, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['Phase', 'Behaviour', 'Start Time', 'End Time', 'Duration (s)'])
                for event in self.behaviour_events:
                    writer.writerow([
                        event.get('phase', 'Default Phase'),
                        event.get('behaviour', ''),
                        self.format_time(event.get('start_time', 0.0)),
                        self.format_time(event.get('end_time', 0.0)),
                        round(float(event.get('duration', 0.0)), 3)
                    ])
        except Exception as e:
            logging.error(f'Error writing CSV: {e}')

    def export_excel(self):
        try:
            options = QFileDialog.Options()
            file_name, _ = QFileDialog.getSaveFileName(self, 'Export to Excel', '', 'Excel Files (*.xlsx);;All Files (*)', options=options)
            if not file_name:
                return
            if not file_name.lower().endswith('.xlsx'):
                file_name += '.xlsx'
            wb = Workbook()
            try:
                wb.remove(wb.active)
            except Exception:
                pass

            ws_summary = wb.create_sheet('Summary')
            ws_summary.append(['Behaviour', 'Time Spent (s)', 'Count'])
            for behaviour in sorted(self.time_spent.keys()):
                ws_summary.append([
                    behaviour,
                    round(float(self.time_spent.get(behaviour, 0.0)), 3),
                    int(self.frequency_counts.get(behaviour, 0))
                ])

            ws_events = wb.create_sheet('Detailed_Events')
            ws_events.append(['Phase', 'Behaviour', 'Start Time', 'End Time', 'Duration (s)'])
            for event in self.behaviour_events:
                ws_events.append([
                    event.get('phase', 'Default Phase'),
                    event.get('behaviour', ''),
                    self.format_time(event.get('start_time', 0.0)),
                    self.format_time(event.get('end_time', 0.0)),
                    round(float(event.get('duration', 0.0)), 3)
                ])

            self._autosize_excel_sheet(ws_summary)
            self._autosize_excel_sheet(ws_events)
            wb.save(file_name)
            QMessageBox.information(self, 'Exported', f'Exported to {file_name}')
        except Exception as e:
            QMessageBox.warning(self, 'Error', f'Failed to export Excel: {e}')
            logging.error(f'Failed to export Excel: {e}')

    def _autosize_excel_sheet(self, ws):
        try:
            for col_cells in ws.columns:
                max_len = 0
                col_idx = col_cells[0].column
                for cell in col_cells:
                    v = cell.value
                    if v is None:
                        continue
                    max_len = max(max_len, len(str(v)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
        except Exception as e:
            logging.warning(f'Could not autosize Excel columns: {e}')

    def get_all_behaviour_events(self):
        return self.behaviour_events

    def new_session(self):
        try:
            reply = QMessageBox.question(
                self, 'New Session', 'Start a new session? This will clear current data.',
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return
            if self.video_thread and self.video_thread.isRunning():
                self.video_thread.stop()
            self.scoring_timer.stop()
            self.scoring = False
            self.time_spent = {}
            self.frequency_counts = {}
            self.behaviour_events = []
            self.undo_stack = []
            self.redo_stack = []
            self.annotations = []
            self.paused_frame = None
            self.update_behaviour_labels()
            self.update_frequency_controls()
            self.timeline_widget.update_events(self.get_all_behaviour_events())
            self.video_slider.setValue(0)
            self.update_timestamp(0, self.total_duration)
            self.video_widget.image = None
            self.video_widget.update()
            self.feedback_label.setText('New session started. Assign keys and start scoring.')
            self.history_panel.clear()
            logging.info('New session started and previous data reset.')
        except Exception as e:
            QMessageBox.warning(self, 'Error', f'Failed to reset session: {e}')
            logging.error(f'Failed to reset session: {e}')

    def parse_time_to_seconds(self, time_str):
        try:
            if isinstance(time_str, (int, float)):
                return float(time_str)
            hms, _, ms = str(time_str).partition('.')
            h, m, s = [int(x) for x in hms.split(':')]
            ms = int((ms + '000')[:3])
            return h * 3600 + m * 60 + s + ms / 1000.0
        except Exception as e:
            logging.error(f'Error parsing time: {e}')
            return 0.0

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        try:
            for url in event.mimeData().urls():
                file_path = url.toLocalFile()
                if os.path.isfile(file_path):
                    ext = os.path.splitext(file_path)[1].lower()
                    if ext in ['.avi', '.mp4', '.mov', '.mkv']:
                        self.load_video(file_path)
                        break
        except Exception as e:
            logging.error(f'Error handling drop event: {e}')

    def autosave_session(self):
        try:
            if self.behaviour_events:
                autosave_path = 'autosave_rms.csv'
                self.write_csv(autosave_path)
                logging.info(f'Autosaved {len(self.behaviour_events)} events to autosave_rms.csv.')
            else:
                logging.info('Autosave heartbeat — no events to save.')
        except Exception as e:
            logging.error(f'Autosave error: {e}')

    def closeEvent(self, event):
        try:
            if self.video_thread and self.video_thread.isRunning():
                self.video_thread.stop()
        except Exception:
            pass
        event.accept()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = BehaviourScoringApp()
    window.show()
    sys.exit(app.exec_())
